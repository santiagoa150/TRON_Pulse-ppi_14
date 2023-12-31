import os
import io

import requests
import numpy as np
import pandas as pd
from openpyxl import Workbook
import matplotlib.pyplot as plt
from matplotlib.backends.backend_agg import FigureCanvasAgg as FigureCanvas
import qrcode

from rest_framework.decorators import api_view
from django.http import HttpResponse
from users.authentication import authenticate
from .constants import TronApiConstants
from .utils import TronApiUtils
from core.utils import ApiUtils
from django.shortcuts import render
from django.http import JsonResponse
import requests



def validate_address_util(address) -> bool:
    """
    Valida una dirección en la red utilizando un servicio externo.

    Esta función se encarga de validar si una dirección proporcionada\
    es válida en la red utilizando un servicio externo.
    Realiza una solicitud POST al servicio externo con la dirección como\
    parámetro y retorna un valor booleano que indica
    si la dirección es válida o no.

    Args:
        address (str): La dirección que se desea validar.

    Returns:
        bool: True si la dirección es válida, False si no lo es.
    """

    # Realiza una solicitud POST a la API de Tron
    service_response = requests.post(TronApiConstants.VALIDATE_ADDRESS_URL.value, json={'address': address})

    # Extrae el resultado de la respuesta del servicio
    return service_response.json()['result']


def validate_contract_util(address) -> bool:
    """
        Valida la existencia de un contrato utilizando una API externa.

        Esta función envía una solicitud POST a una API externa con la dirección del contrato.
        Si la respuesta contiene un campo "abi", se considera que el contrato existe.

        Argumentos:
        - address (str): Dirección del contrato a validar.

        Retorna:
        - bool: True si el contrato existe; False en caso contrario.
        """
    # Envía la solicitud a la API externa para obtener información sobre el contrato
    service_response = requests.post(TronApiConstants.GET_CONTRACT_URL.value, json={"value": address, "visible": True})
    response = False

    # Verifica si la respuesta de la API contiene el campo "abi" para determinar la existencia del contrato
    if "abi" in service_response.json():
        response = True
    return response


def get_trx_balance_util(address) -> int:
    """
    Obtiene el saldo TRX de una dirección específica en la red Tron.

    Esta función realiza una solicitud a la API de Tron para\
    obtener el saldo de TRX de la dirección especificada.
    El saldo se obtiene en unidades SUN y se convierte a TRX\
    utilizando una constante de conversión.

    Args:
        address (str): La dirección de TRX de la que se desea obtener el saldo.

    Returns:
        int: El saldo TRX de la dirección.
    """

    # Construye la URL para obtener el saldo de TRX de la dirección
    url = TronApiConstants.GET_BALANCE_URL.value.replace(TronApiConstants.REPLACE_ADDRESS_PARAM.value, address)

    # Realiza una solicitud a la API de Tron y obtiene la respuesta en formato JSON
    raw_data = requests.get(url).json()['data']

    # Extrae el saldo en unidades SUN de la respuesta
    balance = 0

    try:
        raw_balance = raw_data[0]['balance']

        # Convierte el saldo de SUN a TRX utilizando una constante de conversión
        balance = raw_balance / TronApiConstants.SUN_TO_TRX.value
    except:
        pass

    return balance


def get_trm(symbol: str) -> int:
    """
    Obtiene la tasa de cambio en USD para una criptomoneda específica.

    Esta función realiza una solicitud a una fuente de datos externa\
    para obtener la tasa de cambio en USD de una criptomoneda.
    El valor se utiliza para convertir entre la criptomoneda y USD.

    Args:
        symbol (str): El símbolo de la criptomoneda para la que se desea obtener la tasa de cambio.

    Returns:
        int: El valor de la tasa de cambio en USD para la criptomoneda.

    Note:
        Esta función depende de una clave de API válida definida en\
        la variable de entorno 'COINMARKET_API_KEY'.
    """

    # Construye la URL para obtener la tasa de cambio de la criptomoneda en USD
    url = TronApiConstants.GET_BALANCE_URL.GET_TRM_URL.value
    params = {
        'CMC_PRO_API_KEY': os.getenv('COINMARKET_API_KEY'),
        'symbol': symbol,
        'convert': 'USD',
    }

    # Realiza una solicitud a Coin Market y obtiene la respuesta en formato JSON
    raw_data = requests.get(url, params=params).json()

    # Extrae el valor de la tasa de cambio en USD para la criptomoneda especificada
    return raw_data['data'][symbol]['quote']['USD']['price']


@api_view(["GET"])
def validate_address(req, address):
    """
    Valida una dirección en la red.

    Esta vista se encarga de validar si una dirección proporcionada \
    es válida en la red. Utiliza una función de utilidad
    llamada 'validate_address_util' para realizar la validación y \
    devuelve un resultado en formato JSON indicando si la
    dirección es válida o no.

    Args:
        req (HttpRequest): La solicitud HTTP que puede contener parámetros o información adicional.
        address (str): La dirección que se desea validar.

    Returns:
        JsonResponse: Una respuesta JSON que contiene un indicador de si la dirección es válida o no.

    Note:
        La validación se realiza utilizando la función 'validate_address_util', \
        y esta vista no requiere autenticación.
    """

    # Utiliza la función de utilidad 'validate_address_util' para validar la dirección
    return ApiUtils.build_generic_response({'isValid': validate_address_util(address)})


@api_view(["GET"])
def get_trx_balance(req, address):
    """
    Obtiene el saldo TRX de una dirección específica en la red Tron.

    Esta función utiliza una función de utilidad para obtener el saldo\
    de TRX de la dirección proporcionada.Opcionalmente, puede calcular\
    el saldo en USD si se proporciona un token válido y se establece el\
    parámetro 'requiresUSD' en 'true'.

    Args:
        req (HttpRequest): La solicitud HTTP que contiene los parámetros de la solicitud.
        address (str): La dirección de TRX de la que se desea obtener el saldo.

    Returns:
        JsonResponse: Una respuesta JSON que contiene el saldo TRX y, si es necesario, el saldo en USD.

    Raises:
        UnauthorizedResponse: Si la autenticación del token es requerida y no es válida.
    """

    # Obtiene el saldo de TRX utilizando una función de utilidad
    balance = get_trx_balance_util(address)
    response = {'balance': balance}

    # Verifica si se requieren valores en USD y si se proporciona un token válido
    if req.GET.get('requiresUSD') == 'true':
        # Verifica la autenticación del token
        if not authenticate(req.GET.get('token')):
            return ApiUtils.build_unauthorized_response()

        # Obtiene la tasa de cambio en USD para TRX y calcula el saldo en USD
        trm = get_trm('TRX')
        usd_balance = balance * trm
        response['USDBalance'] = usd_balance

    # Retorna una respuesta JSON que incluye el saldo TRX y, si es necesario, el saldo en USD
    return ApiUtils.build_generic_response(response)


@api_view(["GET"])
def get_trx_transactions(req, address):
    """
    Obtiene transacciones de TRX relacionadas con una dirección específica en la red Tron.

    Esta función realiza una solicitud a la API de Tron para obtener\
    transacciones de TRX relacionadas con una dirección
    específica. Puede filtrar las transacciones por rango de tiempo y\
    calcular estadísticas de las transacciones.

    Args:
        req (HttpRequest): La solicitud HTTP que contiene los parámetros de la solicitud.
        address (str): La dirección de TRX de la que se desean obtener las transacciones.

    Returns:
        JsonResponse: Una respuesta JSON que contiene las transacciones\
        de TRX y estadísticas relacionadas con la dirección.

    Raises:
        UnauthorizedResponse: Si la autenticación del token es requerida y no es válida.

    Note:
        Si se requieren valores en USD, se debe proporcionar un token\
        válido y establecer el parámetro 'requiresUSD' en 'true'.
    """

    url = TronApiConstants.GET_TRANSACTIONS_URL.value.replace(TronApiConstants.REPLACE_ADDRESS_PARAM.value, address)
    params = {}

    # Verifica si se proporciona un rango de tiempo para filtrar las transacciones
    if req.GET.get('startTimestamp'):
        params['min_timestamp'] = req.GET.get('startTimestamp')

    if req.GET.get('finalTimestamp'):
        params['max_timestamp'] = req.GET.get('finalTimestamp')

    # Inicializa listas para almacenar las transacciones y estadísticas
    transactions = []
    statistics = {
        'average': 0,
        'med': 0,
        'maximum': 0,
        'minimum': 0,
        'sum': 0,
    }

    # Realiza una solicitud a la API de Tron y obtiene la respuesta en formato JSON
    raw_response = requests.get(url, params=params).json()['data']
    if raw_response:
        # Mapea y procesa la respuesta para obtener datos relevantes
        data = TronApiUtils.map_get_trx_transactions_response(raw_response)
        data['amount'] = data['amount'] / TronApiConstants.SUN_TO_TRX.value
        data['from'] = TronApiUtils.hex_address_to_base58(data['from'])
        data['to'] = TronApiUtils.hex_address_to_base58(data['to'])
        data['type'] = np.where(data['to'] == address, 'Entrada', 'Salida')

        # Calcula estadísticas de las transacciones
        inputs = data.loc[data['type'] == 'Entrada', 'amount'].sum()
        outputs = data.loc[data['type'] == 'Salida', 'amount'].sum()
        statistics['average'] = data['amount'].mean()
        statistics['med'] = data['amount'].median()
        statistics['maximum'] = data['amount'].max()
        statistics['minimum'] = data['amount'].min()
        statistics['sum'] = inputs - outputs

        # Verifica si se requieren valores en USD y si se proporciona un token válido
        if req.GET.get('requiresUSD') == 'true':
            # Verifica la autenticación del token
            if not authenticate(req.GET.get('token')):
                return ApiUtils.build_unauthorized_response()

            # Obtiene la tasa de cambio en USD para TRX
            trm = get_trm('TRX')
            data['USDAmount'] = data['amount'] * trm

        # Convierte los datos de pandas en una lista para la respuesta.
        transactions = data.values.tolist()
    return ApiUtils.build_generic_response({'transactions': transactions, 'statistics': statistics})


@api_view(["GET"])
def get_history_blocks(req, quantity: int):
    """
    Obtiene un historial de bloques de la red Tron.

    Esta función realiza una solicitud a la API de Tron para\
    obtener un historial de bloques de la red Tron.
    Los bloques se pueden filtrar por cantidad.

    Args:
        req (HttpRequest): La solicitud HTTP que contiene los parámetros de la solicitud.
        quantity (int): La cantidad de bloques a obtener en el historial.

    Returns:
        JsonResponse: Una respuesta JSON que contiene el historial de bloques solicitado.

    Note:
        En caso de que no se puedan obtener datos de bloques,\
        la respuesta será una lista vacía en formato JSON.
    """

    url = TronApiConstants.GET_BLOCK_HISTORY_URL.value
    params = {'num': quantity}

    # Se consultan los bloques en la red de tron.
    raw_data = requests.get(url, params=params).json()
    blocks = []
    try:
        # Mapear los bloques encontrados.
        data = TronApiUtils.map_get_block_history_response(raw_data['block'])

        # Se retorna el inverso de los datos para mostrar\
        # el más actual primero.
        blocks = data.values.tolist()[::-1]
    except:
        # En caso de error o falta de datos, se maneja\
        # la excepción y se devuelve una lista vacía.
        pass
    return ApiUtils.build_generic_response({'blocks': blocks})


@api_view(["GET"])
def generate_qr_code(request, address):
    """
    Genera una imagen de código QR para una dirección dada y la devuelve como una respuesta HTTP.

    Args:
    - request: Objeto de solicitud de Django.
    - direccion (str): La dirección para codificar en el código QR.

    Returns:
    - HttpResponse: Respuesta HTTP que contiene la imagen del código QR en formato PNG.
    """

    # Crear la instancia del código QR
    qr = qrcode.QRCode(
        version=1,
        error_correction=qrcode.constants.ERROR_CORRECT_L,
        box_size=10,
        border=4,
    )

    # Agregar los datos de la dirección al código QR
    qr.add_data(address)
    qr.make(fit=True)

    # Crear una figura de Matplotlib
    fig, ax = plt.subplots()
    img = qr.make_image(fill_color="black", back_color="white")
    ax.imshow(img)

    # Configurar la figura y el eje, se ocultan
    ax.axis('off')

    # Convertir la figura a un flujo de bytes
    canvas = FigureCanvas(fig)
    buf = io.BytesIO()
    canvas.print_png(buf)
    plt.close(fig)

    # Crear una respuesta HTTP con la imagen del código QR
    response = HttpResponse(buf.getvalue(), content_type='image/png')
    return response


@api_view(["GET"])
def get_block_transactions(req, block: int):
    """
    Obtiene las transacciones de un bloque específico en la red Tron.

    Esta función realiza una solicitud a la API de Tron para\
    obtener las transacciones de un bloque determinado.
    También puede calcular la tarifa del minero en TRX y, opcionalmente,\
    en USD si se proporciona un token válido.

    Args:
        req (HttpRequest): La solicitud HTTP que contiene los parámetros de la solicitud.
        block (int): El número de bloque del que se desean obtener las transacciones.

    Returns:
        JsonResponse: Una respuesta JSON que contiene las transacciones del bloque solicitado.

    Raises:
        UnauthorizedResponse: Si la autenticación del token es requerida y no es válida.

    Note:
        Para obtener las transacciones en USD, se debe proporcionar\
        un token válido y establecer el parámetro 'requiresUSD' en 'true'.
    """

    url = TronApiConstants.GET_BLOCK_TRANSACTIONS_URL.value
    params = {'num': block}
    # Consultar las transacciones del bloque
    raw_data = requests.get(url, params=params).json()

    # Mapear las transacciones encontradas
    data = TronApiUtils.map_get_block_transactions(raw_data)
    data['minerFee'] = data['minerFee'] / TronApiConstants.SUN_TO_TRX.value

    if req.GET.get('requiresUSD') == 'true':
        # Verificar la autenticación del token
        if not authenticate(req.GET.get('token')):
            return ApiUtils.build_unauthorized_response()

        # Obtener la tasa de cambio en USD para TRX (si es necesario)
        trm = get_trm('TRX')
        data['USDMinerFee'] = data['minerFee'] * trm

    return ApiUtils.build_generic_response({'transactions': data.values.tolist()})

def get_contract_transactions(request, contract_address):
    """
    Obtiene la información de transacciones asociadas a un contrato en la red Tron.

    Parameters:
        - request: La solicitud HTTP recibida.
        - contract_address (str): La dirección del contrato Tron del cual se obtendrán las transacciones.

    Returns:
        - JsonResponse: Una respuesta JSON que incluye la información de transacciones y otros datos relevantes.

    Raises:
        - No specific exceptions raised.

    """

    # Construir la URL para obtener transacciones asociadas al contrato desde la API de Tron
    url = f"https://api.shasta.trongrid.io/v1/contracts/{contract_address}/transactions"

    # Realizar una solicitud GET a la API de Tron
    response = requests.get(url)

    # Convertir la respuesta a formato JSON
    data = response.json()

    # Extraer la información necesaria del contrato
    contract_address = data.get("contract_address", "")
    bytecode = data.get("bytecode", "")

    # Devolver una respuesta JSON con la información obtenida
    return JsonResponse({"result": "success", 'response': data})


@api_view(["GET"])
def export_to_excel(request, address):
    """
    Exporta transacciones relacionadas con la dirección Tron a un archivo Excel.

    Parámetros:
        - request: La solicitud HTTP recibida.
        - address (str): La dirección Tron para la cual se obtendrán las transacciones.

    Retorna:
        - HttpResponse: Una respuesta HTTP con el archivo Excel adjunto o una respuesta JSON vacía.

    Excepciones:
        - No se generan excepciones específicas.

    """
    
    # Obtener las categorías seleccionadas de los parámetros de la consulta
    selected_categories = request.query_params.get('categories', '').split(',')

    # Construir la URL para obtener transacciones desde la API de Tron
    url = TronApiConstants.GET_TRANSACTIONS_URL.value.replace(TronApiConstants.REPLACE_ADDRESS_PARAM.value, address)
    params = {}

    # Obtener datos de transacciones desde la API de Tron
    raw_response = requests.get(url, params=params).json()['data']

    if raw_response:
        # Mapear y procesar la respuesta para obtener datos relevantes
        data = TronApiUtils.map_get_trx_transactions_response(raw_response)
        data['amount'] = data['amount'] / TronApiConstants.SUN_TO_TRX.value
        data['from'] = TronApiUtils.hex_address_to_base58(data['from'])
        data['to'] = TronApiUtils.hex_address_to_base58(data['to'])
        data['type'] = np.where(data['to'] == address, 'Entrada', 'Salida')

        # Filtrar datos según las categorías seleccionadas
        data = data[selected_categories]

        # Crear un DataFrame de pandas
        df = pd.DataFrame(data)

        # Crear un libro de trabajo de Excel y agregar el DataFrame como una hoja
        wb = Workbook()
        ws = wb.active

        # Escribir las columnas del DataFrame como encabezados
        for col_num, header in enumerate(df.columns, 1):
            ws.cell(row=1, column=col_num, value=header)

        # Escribir los valores del DataFrame en la hoja
        for r_idx, row in enumerate(df.iterrows(), 2):
            for c_idx, value in enumerate(row[1], 1):
                ws.cell(row=r_idx, column=c_idx, value=value)

        # Crear la respuesta HTTP
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename=tron_pulse_{address}.xlsx'

        # Guardar el libro de trabajo y escribir la respuesta
        wb.save(response)

        return response
    else:
        # Si no hay datos, devolver una respuesta JSON vacía
        return ApiUtils.build_generic_response({'transactions': [], 'statistics': {}})
